"""App-owned openpyxl workbook export."""
from __future__ import annotations
from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from .conversions import format_split, watts_to_split_seconds, two_k_seconds_to_watts

HEAD=PatternFill("solid", fgColor="1F4E78")
def _sheet(wb,title,headers,rows):
    ws=wb.create_sheet(title); ws.append(headers)
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=HEAD; c.alignment=Alignment(wrap_text=True)
    for r in rows: ws.append(r)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(45,max(12,max(len(str(x.value or "")) for x in col)+2))
    return ws

def build_workbook(profile: dict, plan: dict, source_rows: list[list[str]] | None = None) -> bytes:
    wb=Workbook(); ws=wb.active; ws.title="START HERE"
    ws.append(["Rowing Plan Generator"]); ws["A1"].font=Font(size=16,bold=True,color="FFFFFF"); ws["A1"].fill=HEAD
    for row in [["Athlete",profile["athlete"]["display_name"]],["Season",f"{profile['season']['start_date']} to {profile['season']['end_date']}"],["Generated",plan["generated_at"]],["Intensity",plan["intensity_profile"][0]["method"]],["Disclaimer","Planning support only; not medical advice. Review with a qualified coach/clinician as appropriate."],["Assumptions","Measured values, coach inputs, and provisional methods are clearly identified in following sheets."]]: ws.append(row)
    tests=profile.get("tests",{}); raw=plan["power_profile"].get("raw_tests",{})
    two_k = tests.get("erg_2k_seconds")
    two_k_display = f"{int(two_k // 60)}:{int(two_k % 60):02d}" if two_k else None
    _sheet(wb,"ATHLETE PROFILE",["Field","Value"],[["Goals",", ".join(x["goal_type"] for x in profile.get("goals",[]))],["2k time (min:sec)",two_k_display],["Resting HR",tests.get("resting_hr")],["Max HR",tests.get("max_hr")]]+[[f"{k} watts",v.get("value_watts")] for k,v in raw.items() if isinstance(v,dict)])
    pp=plan["power_profile"]
    pp_rows=[["Profile status",pp["status"]],["Testing block",pp.get("testing_block_label")],["Algorithm",pp["algorithm_version"]],["Current ratios","Descriptive relationships among your own test results."],["Limitation","No predicted 2k, threshold, or population weakness label is generated."]]
    pp_rows += [[k,v] for k,v in pp["metrics"].items()]
    pp_rows += [[f"Anchor: {a['name']}",f"{a['target_watts_low']:.1f}–{a['target_watts_high']:.1f} W; {a['formula']}"] for a in pp["anchors"]]
    pp_rows += [["Plan impact",x] for x in pp.get("plan_impacts",[])] + [["Warning",w] for w in pp["warnings"]]
    _sheet(wb,"POWER PROFILE",["Item","Value"],pp_rows)
    current_tests=pp.get("raw_tests",{})
    test_rows=[]
    for protocol,test in current_tests.items():
        watts=test.get("peak_watts",test.get("average_watts"))
        if watts is None and test.get("time_seconds"): watts=two_k_seconds_to_watts(test["time_seconds"])
        test_rows.append([protocol,test.get("test_date"),test.get("time_seconds"),watts,format_split(watts_to_split_seconds(watts)) if watts else None,test.get("average_spm") or test.get("peak_spm"),test.get("average_hr") or test.get("peak_hr"),test.get("notes")])
    _sheet(wb,"CURRENT TESTING BLOCK",["Test","Date","2k time seconds","Watts","Split","Rate","HR","Notes"],test_rows)
    trend_headers=["Block","Date","2k time","2k watts","7-stroke peak","60-sec watts","20-sec watts","30R20 watts","Peak/2k","60sec/2k","60sec/peak","2k watts % change"]
    _sheet(wb,"POWER PROFILE TRENDS",trend_headers,[[r.get("label"),r.get("date"),r.get("two_k_time_seconds"),r.get("two_k"),r.get("seven_stroke_peak"),r.get("sixty_second"),r.get("twenty_second"),r.get("thirty_min_rate_capped"),r.get("peak_to_2k_ratio"),r.get("sixty_to_2k_ratio"),r.get("sixty_to_peak_ratio"),r.get("two_k_pct_change")] for r in pp.get("longitudinal",{}).get("rows",[])])
    _sheet(wb,"TRAINING BANDS",["Band","Domain","HR low","HR high","Watts low","Watts high","Rate","Method","Confidence","Assumptions"],[[b["name"],b["domain"],b.get("hr_low"),b.get("hr_high"),b.get("watts_low"),b.get("watts_high"),f"{b.get('spm_low')}–{b.get('spm_high')}",b["method"],b["confidence"],"; ".join(b.get("assumptions") or [])] for b in plan["intensity_profile"]])
    _sheet(wb,"SEASON OVERVIEW",["Date","Phase","Race event"],[[s["date"],s["phase"],s.get("race_event")] for s in plan.get("phases",[])])
    heads=["Week","Date","Day","Type of row","Coached","HR range","Description of workout","Rating (spm)","Total time (min)"]
    rows=[]
    for s in plan["sessions"]:
        rows.append([date.fromisoformat(s["date"]).isocalendar().week,date.fromisoformat(s["date"]),s["day"],s["band"],"Yes" if s.get("coached") else "No",s.get("hr_range"),s.get("description",s.get("title")),s.get("rating"),s.get("total_cardio_minutes",0)])
    _sheet(wb,"DAILY SCHEDULE",heads,rows)
    _sheet(wb,"WEEKLY TOTALS",["Week","Cardio min","Rowing min","Strength sessions","Quality sessions"],[[x["week"],x["cardio_minutes"],x["rowing_minutes"],x["strength_sessions"],x["quality_sessions"]] for x in plan["weekly_totals"]])
    _sheet(wb,"SESSION LIBRARY",["Session ID","Rights","Source basis"],[[s["session_id"],"Original app template",", ".join(s.get("source_basis_ids",[]))] for s in plan["sessions"] if s.get("source_basis_ids")])
    _sheet(wb,"RACE PLAN",["Event","Start","End","Priority","Boat","Distance","Expected races","Notes"],[[r["event_name"],date.fromisoformat(r["start_date"]),date.fromisoformat(r["end_date"]),r["priority"],r.get("boat_class"),r.get("distance_m"),r["expected_races"],r.get("notes")] for r in profile.get("races",[])])
    _sheet(wb,"WEEKLY LOG",["Week","Completed minutes","Average HR","Average watts/split","Average rate","RPE","Sleep/recovery note","Coach comments"],[])
    _sheet(wb,"SOURCES",["Source ID","Citation","URL","Access/license","Notes"],source_rows or [])
    out=BytesIO(); wb.save(out); return out.getvalue()
